import streamlit as st
import pandas as pd
from datetime import datetime, date
import io

# Importation d'openpyxl pour la mise en page Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =========================================================
# FONCTION GENERATION EXCEL FORMAT A4 PORTRAIT
# =========================================================
def generate_excel_synthesis(df_data, titre_periode):
    """Génère un fichier Excel parfaitement mis en page au format A4 Portrait pour impression."""
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Synthèse Béton"

    # --- 1. CONFIGURATION D'IMPRESSION A4 PORTRAIT ---
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    # Marges réduites
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # --- 2. PALETTE DE COULEURS ET STYLES ---
    color_primary = "1F4E79"    # Bleu LPEE / Marine
    color_th = "2D572C"         # Vert/Gris foncé entête
    color_kpi = "F2F4F7"        # Fond clair KPI

    font_title = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    font_section = Font(name="Arial", size=11, bold=True, color=color_primary)
    font_bold = Font(name="Arial", size=9, bold=True)
    font_normal = Font(name="Arial", size=9)
    font_th = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    font_kpi_val = Font(name="Arial", size=12, bold=True, color=color_primary)

    fill_title = PatternFill(start_color=color_primary, end_color=color_primary, fill_type="solid")
    fill_th = PatternFill(start_color=color_th, end_color=color_th, fill_type="solid")
    fill_kpi = PatternFill(start_color=color_kpi, end_color=color_kpi, fill_type="solid")

    thin_side = Side(style='thin', color='D9D9D9')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    total_border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))
    box_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Détermination du nombre de colonnes
    nb_cols = max(len(df_data.columns), 6)
    last_col = get_column_letter(nb_cols)

    # --- 3. BANNIÈRE EN-TÊTE LPEE ---
    ws.merge_cells(f"A1:{last_col}2")
    cell_title = ws["A1"]
    cell_title.value = "LABORATOIRE PUBLIC D'ESSAIS ET D'ÉTUDES (LPEE) - CTR-CSB\nRAPPORT DE SYNTHÈSE DU BÉTONNAGE"
    cell_title.font = font_title
    cell_title.fill = fill_title
    cell_title.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # --- 4. BLOC INFOS CLIENT & PROJET ---
    ws["A4"] = "CLIENT :"
    ws["A4"].font = font_bold
    ws["B4"] = "TGCC"
    ws["B4"].font = font_normal

    ws["D4"] = "PROJET :"
    ws["D4"].font = font_bold
    ws["E4"] = "LGV CASA SUD"
    ws["E4"].font = font_normal

    ws["A5"] = "PÉRIODE :"
    ws["A5"].font = font_bold
    ws["B5"] = titre_periode
    ws["B5"].font = font_normal

    ws["D5"] = "DATE ÉDITION :"
    ws["D5"].font = font_bold
    ws["E5"] = datetime.now().strftime("%d/%m/%Y")
    ws["E5"].font = font_normal

    # --- 5. RÉSUMÉ DES KPIs ---
    row_idx = 7
    ws.merge_cells(f"A{row_idx}:{last_col}{row_idx}")
    ws[f"A{row_idx}"] = "📊 RÉSUMÉ GLOBAL"
    ws[f"A{row_idx}"].font = font_section
    row_idx += 1

    vol_tot = df_data["Quantité (m³)"].sum() if "Quantité (m³)" in df_data.columns else 0
    nb_coulages = len(df_data)
    nb_eprouvettes = df_data["Nb Éprouvettes"].sum() if "Nb Éprouvettes" in df_data.columns else 0

    kpi_items = [
        ("Volume Total Béton", f"{vol_tot:.1f} m³"),
        ("Nombre de Coulages / BL", f"{nb_coulages}"),
        ("Total Éprouvettes", f"{int(nb_eprouvettes)}")
    ]

    for i, (lbl, val) in enumerate(kpi_items):
        col_s = i * 2 + 1
        c1, c2 = get_column_letter(col_s), get_column_letter(col_s + 1)
        
        ws.merge_cells(f"{c1}{row_idx}:{c2}{row_idx}")
        ws[f"{c1}{row_idx}"] = lbl
        ws[f"{c1}{row_idx}"].font = font_bold
        ws[f"{c1}{row_idx}"].fill = fill_kpi
        ws[f"{c1}{row_idx}"].alignment = Alignment(horizontal="center")

        ws.merge_cells(f"{c1}{row_idx+1}:{c2}{row_idx+1}")
        ws[f"{c1}{row_idx+1}"] = val
        ws[f"{c1}{row_idx+1}"].font = font_kpi_val
        ws[f"{c1}{row_idx+1}"].fill = fill_kpi
        ws[f"{c1}{row_idx+1}"].alignment = Alignment(horizontal="center")

    row_idx += 3

    # --- 6. TABLEAU DES DONNÉES ---
    ws.merge_cells(f"A{row_idx}:{last_col}{row_idx}")
    ws[f"A{row_idx}"] = "📋 DÉTAIL DES CONTRÔLES"
    ws[f"A{row_idx}"].font = font_section
    row_idx += 1

    headers = list(df_data.columns)
    for col_num, h_name in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_num)
        cell.value = str(h_name)
        cell.font = font_th
        cell.fill = fill_th
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[row_idx].height = 24
    row_idx += 1

    start_data_row = row_idx
    for row_data in df_data.itertuples(index=False):
        for col_num, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.value = val
            cell.font = font_normal
            cell.border = thin_border
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_idx].height = 20
        row_idx += 1

    # Ligne de TOTAL
    end_data_row = row_idx - 1
    total_cell = ws.cell(row=row_idx, column=1)
    total_cell.value = "TOTAL"
    total_cell.font = font_bold
    total_cell.border = total_border

    for col_num in range(1, len(headers) + 1):
        c = ws.cell(row=row_idx, column=col_num)
        c.border = total_border
        c.font = font_bold
        col_name = headers[col_num - 1]
        col_ltr = get_column_letter(col_num)
        if col_name == "Quantité (m³)":
            c.value = f"=SUM({col_ltr}{start_data_row}:{col_ltr}{end_data_row})"
            c.number_format = '0.0 "m³"'
            c.alignment = Alignment(horizontal="right")
        elif col_name == "Nb Éprouvettes":
            c.value = f"=SUM({col_ltr}{start_data_row}:{col_ltr}{end_data_row})"
            c.alignment = Alignment(horizontal="right")

    row_idx += 3

    # --- 7. PIED DE PAGE : ZONE DE SIGNATURES ---
    ws.merge_cells(f"A{row_idx}:C{row_idx}")
    ws[f"A{row_idx}"] = "Responsables d'essai"
    ws[f"A{row_idx}"].font = font_bold
    ws[f"A{row_idx}"].alignment = Alignment(horizontal="center")

    sig_col_start = get_column_letter(max(nb_cols - 2, 4))
    ws.merge_cells(f"{sig_col_start}{row_idx}:{last_col}{row_idx}")
    ws[f"{sig_col_start}{row_idx}"] = "Chef du Laboratoire"
    ws[f"{sig_col_start}{row_idx}"].font = font_bold
    ws[f"{sig_col_start}{row_idx}"].alignment = Alignment(horizontal="center")

    row_idx += 1
    # Zones de visa / signature
    ws.merge_cells(f"A{row_idx}:C{row_idx+3}")
    ws[f"A{row_idx}"] = "\n\nVisa & Signature :"
    ws[f"A{row_idx}"].font = font_normal
    ws[f"A{row_idx}"].alignment = Alignment(horizontal="left", vertical="top")

    ws.merge_cells(f"{sig_col_start}{row_idx}:{last_col}{row_idx+3}")
    ws[f"{sig_col_start}{row_idx}"] = "\n\nVisa & Signature :"
    ws[f"{sig_col_start}{row_idx}"].font = font_normal
    ws[f"{sig_col_start}{row_idx}"].alignment = Alignment(horizontal="left", vertical="top")

    # Ajustement automatique des largeurs de colonnes
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.row not in [1, 2, 7, row_idx-4] and cell.value:
                val_str = str(cell.value)
                if len(val_str) < 35:
                    max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)

    wb.save(output)
    output.seek(0)
    return output.getvalue()


# =========================================================
# VUE PRINCIPALE STREAMLIT
# =========================================================
def show(supabase):
    st.title("📊 Récapitulatif et Synthèse du Bétonnage")
    
    tab_journalier, tab_mensuel = st.tabs(["📅 Bilan Journalier", "📅 Bilan Mensuel"])
    
    # ---------------------------------------------------------
    # 1. BILAN JOURNALIER
    # ---------------------------------------------------------
    with tab_journalier:
        st.markdown("### Filtrage par jour et par classe de béton")
        
        col1, col2 = st.columns(2)
        with col1:
            selected_date = st.date_input("Sélectionnez une date :", value=date.today())
        with col2:
            selected_class = st.selectbox(
                "Filtrer par classe de béton :", 
                ["Toutes", "C25/30", "C30/37", "C35/45", "C40/50", "C45/55"]
            )
            
        try:
            res = supabase.table("suivi_betonnage").select("*").eq("date_livraison", str(selected_date)).execute()
            data = res.data if res else []
            
            if data:
                df = pd.DataFrame(data)
                
                if selected_class != "Toutes":
                    df = df[df["classe_beton"] == selected_class]
                
                if df.empty:
                    st.info("Aucun coulage enregistré pour les critères sélectionnés.")
                else:
                    # Calcul de la Durée de transport
                    if "heure_fin_coulage" in df.columns and "heure_arrivee" in df.columns:
                        def calc_duree(row):
                            try:
                                h_fin = datetime.strptime(str(row["heure_fin_coulage"]), "%H:%M")
                                h_arr = datetime.strptime(str(row["heure_arrivee"]), "%H:%M")
                                return f"{int((h_arr - h_fin).total_seconds() / 60)} min"
                            except:
                                return "-"
                        df["Durée de transport"] = df.apply(calc_duree, axis=1)

                    # Suppression des colonnes non désirées
                    cols_drop = [c for c in ["id", "created_at", "created", "heure_fin_coulage", "client", "centrale_beton"] if c in df.columns]
                    df = df.drop(columns=cols_drop)

                    # Réorganisation des colonnes (Heure d'arrivée après Date, Météo à la fin)
                    cols = list(df.columns)
                    if "date_livraison" in cols and "heure_arrivee" in cols:
                        cols.remove("heure_arrivee")
                        cols.insert(cols.index("date_livraison") + 1, "heure_arrivee")
                    if "meteo" in cols:
                        cols.remove("meteo")
                        cols.append("meteo")
                    df = df[cols]

                    # Renommage
                    renames = {
                        "date_livraison": "Date Livraison", "heure_arrivee": "Heure d'arrivée",
                        "bl_num": "N° BL", "ouvrage": "Ouvrage", "quantite_m3": "Quantité (m³)",
                        "classe_beton": "Classe", "temperature": "Temp. Béton",
                        "temperature_ambiante": "Temp. Ambiante", "affaissement": "Affaissement",
                        "prelevement": "Prélèvement", "nb_eprouvettes": "Nb Éprouvettes",
                        "observations": "Observations", "technicien": "Technicien", "meteo": "Météo"
                    }
                    df_display = df.rename(columns=renames)

                    # KPIs Visuels
                    st.markdown("---")
                    k1, k2, k3, k4 = st.columns(4)
                    k1.metric("Volume Total", f"{df_display['Quantité (m³)'].sum():.1f} m³")
                    k2.metric("Nombre de Coulages", len(df_display))
                    k3.metric("Affaissement Moyen", f"{df_display['Affaissement'].mean():.0f} mm")
                    k4.metric("Éprouvettes Prélevées", int(df_display['Nb Éprouvettes'].sum()))
                    
                    st.markdown("---")
                    
                    # 📥 Bouton de Téléchargement Excel
                    excel_file = generate_excel_synthesis(df_display, f"Journée du {selected_date.strftime('%d/%m/%Y')}")
                    st.download_button(
                        label="📥 Télécharger la Synthèse Excel (Format A4)",
                        data=excel_file,
                        file_name=f"Synthese_Beton_{selected_date}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                    st.dataframe(df_display, use_container_width=True)
            else:
                st.info("Aucun coulage enregistré pour les critères sélectionnés.")
                
        except Exception as e:
            st.error(f"Erreur de chargement : {e}")

    # ---------------------------------------------------------
    # 2. BILAN MENSUEL
    # ---------------------------------------------------------
    with tab_mensuel:
        st.markdown("### Bilan mensuel global")
        
        col_m1, col_m2 = st.columns(2)
        with col_m1:
            annee = date.today().year
            mois_liste = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
            mois_selected = st.selectbox("Sélectionnez le mois :", mois_liste, index=date.today().month - 1)
            mois_num = mois_liste.index(mois_selected) + 1
            
        with col_m2:
            selected_class_m = st.selectbox(
                "Filtrer par classe de béton (Mensuel) :", 
                ["Toutes", "C25/30", "C30/37", "C35/45", "C40/50", "C45/55"],
                key="class_mensuel"
            )
            
        try:
            date_debut = f"{annee}-{mois_num:02d}-01"
            dernier_jour = 31 if mois_num in [1,3,5,7,8,10,12] else (30 if mois_num in [4,6,9,11] else 28)
            date_fin = f"{annee}-{mois_num:02d}-{dernier_jour}"
            
            res_m = supabase.table("suivi_betonnage").select("*").gte("date_livraison", date_debut).lte("date_livraison", date_fin).execute()
            data_m = res_m.data if res_m else []
            
            if data_m:
                df_m = pd.DataFrame(data_m)
                if selected_class_m != "Toutes":
                    df_m = df_m[df_m["classe_beton"] == selected_class_m]
                    
                if df_m.empty:
                    st.info("Aucun coulage enregistré pour ce mois.")
                else:
                    # Traitement identique
                    if "heure_fin_coulage" in df_m.columns and "heure_arrivee" in df_m.columns:
                        def calc_duree(row):
                            try:
                                h_fin = datetime.strptime(str(row["heure_fin_coulage"]), "%H:%M")
                                h_arr = datetime.strptime(str(row["heure_arrivee"]), "%H:%M")
                                return f"{int((h_arr - h_fin).total_seconds() / 60)} min"
                            except:
                                return "-"
                        df_m["Durée de transport"] = df_m.apply(calc_duree, axis=1)

                    cols_drop = [c for c in ["id", "created_at", "created", "heure_fin_coulage", "client", "centrale_beton"] if c in df_m.columns]
                    df_m = df_m.drop(columns=cols_drop)

                    cols_m = list(df_m.columns)
                    if "date_livraison" in cols_m and "heure_arrivee" in cols_m:
                        cols_m.remove("heure_arrivee")
                        cols_m.insert(cols_m.index("date_livraison") + 1, "heure_arrivee")
                    if "meteo" in cols_m:
                        cols_m.remove("meteo")
                        cols_m.append("meteo")
                    df_m = df_m[cols_m]

                    renames = {
                        "date_livraison": "Date Livraison", "heure_arrivee": "Heure d'arrivée",
                        "bl_num": "N° BL", "ouvrage": "Ouvrage", "quantite_m3": "Quantité (m³)",
                        "classe_beton": "Classe", "temperature": "Temp. Béton",
                        "temperature_ambiante": "Temp. Ambiante", "affaissement": "Affaissement",
                        "prelevement": "Prélèvement", "nb_eprouvettes": "Nb Éprouvettes",
                        "observations": "Observations", "technicien": "Technicien", "meteo": "Météo"
                    }
                    df_m_display = df_m.rename(columns=renames)

                    st.markdown("---")
                    m1, m2, m3 = st.columns(3)
                    m1.metric("Volume Cumulé du Mois", f"{df_m_display['Quantité (m³)'].sum():.1f} m³")
                    m2.metric("Nombre Total de BL", len(df_m_display))
                    m3.metric("Total Éprouvettes", int(df_m_display['Nb Éprouvettes'].sum()))
                    
                    st.markdown("---")
                    
                    # 📥 Bouton de Téléchargement Excel Mensuel
                    excel_file_m = generate_excel_synthesis(df_m_display, f"Mois de {mois_selected} {annee}")
                    st.download_button(
                        label="📥 Télécharger la Synthèse Mensuelle Excel (Format A4)",
                        data=excel_file_m,
                        file_name=f"Synthese_Mensuelle_{mois_selected}_{annee}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

                    st.dataframe(df_m_display, use_container_width=True)
            else:
                st.info("Aucun coulage enregistré pour ce mois.")
                
        except Exception as e:
            st.error(f"Erreur de chargement : {e}")
