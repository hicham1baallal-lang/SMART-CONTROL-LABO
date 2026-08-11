import streamlit as st
import pandas as pd
from datetime import datetime, date
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel_recap(df_filtered, filter_title="Synthèse Générale"):
    """
    Génère un fichier Excel parfaitement mis en page au format A4 Paysage
    avec en-tête, pied de page, couleurs, formules et synthèse qualité.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Synthèse Essais Plaque"

    # --- CONFIGURATION IMPRESSION A4 PAYSAGE ---
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # --- EN-TÊTE ET PIED DE PAGE D'IMPRESSION (NATIVE EXCEL A4) ---
    ws.oddHeader.left.text = "&\"Calibri,Bold\"&10LABORATOIRE LPEE - CTR-CSB\nProjet: LGV CASA SUD | Client: TGCC"
    ws.oddHeader.center.text = f"&\"Calibri,Bold\"&13SYNTHÈSE DES ESSAIS À LA PLAQUE\n{filter_title}"
    ws.oddHeader.right.text = "&\"Calibri,Regular\"&9Edité le: &D"

    ws.oddFooter.left.text = "&\"Calibri,Italic\"&9Document Officiel - Contrôle Qualité Terrassement (NF P 94-117-1)"
    ws.oddFooter.center.text = "&\"Calibri,Bold\"&10Page &P sur &N"
    ws.oddFooter.right.text = "&\"Calibri,Regular\"&9Visa Laboratoire: _____________"

    # --- PALETTE DE COULEURS ET STYLES ---
    NAVY_HEADER = "1F4E79"
    BLUE_SUBHEADER = "2F5597"
    ICE_BLUE_BG = "F2F5F9"
    BORDER_COLOR = "D9D9D9"
    GREEN_OK = "E2EFDA"
    TEXT_GREEN = "276A3C"
    ORANGE_WARN = "FFF2CC"
    TEXT_ORANGE = "B25900"

    font_title = Font(name="Calibri", size=14, bold=True, color=NAVY_HEADER)
    font_subtitle = Font(name="Calibri", size=10, italic=True, color="595959")
    font_th = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    
    fill_th = PatternFill(start_color=NAVY_HEADER, end_color=NAVY_HEADER, fill_type="solid")
    fill_zebra = PatternFill(start_color=ICE_BLUE_BG, end_color=ICE_BLUE_BG, fill_type="solid")
    fill_kpi = PatternFill(start_color="EAECEE", end_color="EAECEE", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='thin', color=BORDER_COLOR),
        bottom=Side(style='thin', color=BORDER_COLOR)
    )

    thick_top_bottom = Border(
        left=Side(style='thin', color=BORDER_COLOR),
        right=Side(style='thin', color=BORDER_COLOR),
        top=Side(style='medium', color=NAVY_HEADER),
        bottom=Side(style='double', color=NAVY_HEADER)
    )

    # --- 1. BLOC TITRE DU FICHIER ---
    ws.merge_cells("A1:J1")
    ws["A1"] = "LABORATOIRE LPEE — CENTRE TECHNIQUE RÉGIONAL"
    ws["A1"].font = font_title
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:J2")
    ws["A2"] = f"SYNTHÈSE DES ESSAIS DE PORTANCE À LA PLAQUE — {filter_title.upper()}"
    ws["A2"].font = Font(name="Calibri", size=11, bold=True, color=BLUE_SUBHEADER)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:J3")
    ws["A3"] = "Norme : NF P 94-117-1 | Plaque Ø 600 mm | Formules : EV1 = 112.5/(2*Z1), EV2 = 90/(2*Z2), K = EV2/EV1"
    ws["A3"].font = font_subtitle
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 8

    # --- 2. ENTÊTES DE TABLEAU ---
    headers = [
        "Date Essai", "Technicien", "Couche", "Emplacement", 
        "PK / Profil", "Z1 (mm)", "Z2 (mm)", "EV1 (MPa)", "EV2 (MPa)", "K (EV2/EV1)"
    ]

    ws.row_dimensions[5].height = 25
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=text)
        cell.font = font_th
        cell.fill = fill_th
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # --- 3. REMPLISSAGE DES DONNÉES ---
    start_row = 6
    for r_idx, (_, row) in enumerate(df_filtered.iterrows(), start=start_row):
        ws.row_dimensions[r_idx].height = 20
        is_even = (r_idx % 2 == 0)
        current_fill = fill_zebra if is_even else PatternFill(fill_type=None)

        values = [
            str(row.get("date_essai", "")),
            str(row.get("technicien", "")),
            str(row.get("couche", "")),
            str(row.get("emplacement", "")),
            str(row.get("pk_profil", "")),
            float(row.get("z1", 0.0)),
            float(row.get("z2", 0.0)),
            float(row.get("ev1", 0.0)),
            float(row.get("ev2", 0.0)),
            float(row.get("k", 0.0))
        ]

        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            cell.fill = current_fill

            if c_idx in [1, 2]: # Date, Technicien
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif c_idx in [3, 4, 5]: # Couche, Emplacement, PK
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif c_idx in [6, 7]: # Z1, Z2
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "0.00"
            elif c_idx in [8, 9]: # EV1, EV2
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0.00"
            elif c_idx == 10: # K
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "0.00"
                if val >= 1.5:
                    cell.fill = PatternFill(start_color=GREEN_OK, end_color=GREEN_OK, fill_type="solid")
                    cell.font = Font(name="Calibri", size=10, bold=True, color=TEXT_GREEN)
                else:
                    cell.fill = PatternFill(start_color=ORANGE_WARN, end_color=ORANGE_WARN, fill_type="solid")
                    cell.font = Font(name="Calibri", size=10, bold=True, color=TEXT_ORANGE)

    end_row = start_row + len(df_filtered) - 1

    # --- 4. LIGNE DE MOYENNE AUTOMATIQUE ---
    if len(df_filtered) > 0:
        stat_row = end_row + 1
        ws.row_dimensions[stat_row].height = 22

        ws.merge_cells(start_row=stat_row, start_column=1, end_row=stat_row, end_column=5)
        lbl_cell = ws.cell(row=stat_row, column=1, value="MOYENNE DES ESSAIS")
        lbl_cell.font = Font(name="Calibri", size=10, bold=True, color=NAVY_HEADER)
        lbl_cell.alignment = Alignment(horizontal="right", vertical="center")

        for col_idx in range(1, 6):
            ws.cell(row=stat_row, column=col_idx).border = thick_top_bottom
            ws.cell(row=stat_row, column=col_idx).fill = fill_kpi

        # Formules Excel
        formulas = [
            (6, f"=AVERAGE(F{start_row}:F{end_row})", "0.00"),
            (7, f"=AVERAGE(G{start_row}:G{end_row})", "0.00"),
            (8, f"=AVERAGE(H{start_row}:H{end_row})", "#,##0.00"),
            (9, f"=AVERAGE(I{start_row}:I{end_row})", "#,##0.00"),
            (10, f"=AVERAGE(J{start_row}:J{end_row})", "0.00")
        ]

        for c_idx, form, num_fmt in formulas:
            c = ws.cell(row=stat_row, column=c_idx, value=form)
            c.font = Font(name="Calibri", size=10, bold=True, color=NAVY_HEADER)
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border = thick_top_bottom
            c.fill = fill_kpi
            c.number_format = num_fmt

        # --- 5. SYNTHÈSE QUALITÉ (STYLE SYNTHÈSE BÉTON) ---
        synth_start = stat_row + 2
        ws.cell(row=synth_start, column=1, value="RÉSUMÉ STATISTIQUE QUALITÉ (PAYSAGE A4)").font = Font(name="Calibri", size=11, bold=True, color=NAVY_HEADER)

        summary_headers = ["Indicateur", "EV1 (MPa)", "EV2 (MPa)", "Ratio K (EV2/EV1)"]
        ws.row_dimensions[synth_start+1].height = 22

        for idx, header in enumerate(summary_headers, start=1):
            c = ws.cell(row=synth_start+1, column=idx, value=header)
            c.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color=BLUE_SUBHEADER, end_color=BLUE_SUBHEADER, fill_type="solid")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border

        metrics = [
            ("Valeur Minimale", f"=MIN(H{start_row}:H{end_row})", f"=MIN(I{start_row}:I{end_row})", f"=MIN(J{start_row}:J{end_row})"),
            ("Valeur Maximale", f"=MAX(H{start_row}:H{end_row})", f"=MAX(I{start_row}:I{end_row})", f"=MAX(J{start_row}:J{end_row})"),
            ("Moyenne Générale", f"=AVERAGE(H{start_row}:H{end_row})", f"=AVERAGE(I{start_row}:I{end_row})", f"=AVERAGE(J{start_row}:J{end_row})"),
            ("Nombre total d'essais", f"=COUNT(H{start_row}:H{end_row})", f"=COUNT(I{start_row}:I{end_row})", f"=COUNT(J{start_row}:J{end_row})")
        ]

        for idx, (label, ev1_f, ev2_f, k_f) in enumerate(metrics, start=synth_start+2):
            ws.row_dimensions[idx].height = 19
            c1 = ws.cell(row=idx, column=1, value=label)
            c2 = ws.cell(row=idx, column=2, value=ev1_f)
            c3 = ws.cell(row=idx, column=3, value=ev2_f)
            c4 = ws.cell(row=idx, column=4, value=k_f)
            
            c1.font = Font(name="Calibri", size=10, bold=True)
            c1.border = thin_border
            c1.alignment = Alignment(horizontal="left", vertical="center")
            
            for c, fmt in zip([c2, c3, c4], ["#,##0.00", "#,##0.00", "0.00" if "Nombre" not in label else "0"]):
                c.font = Font(name="Calibri", size=10)
                c.border = thin_border
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = fmt

    # --- AJUSTEMENT DES LARGEURS DE COLONNES POUR L'IMPRESSION A4 ---
    col_widths = {
        'A': 13, 'B': 14, 'C': 18, 'D': 20, 'E': 14,
        'F': 10, 'G': 10, 'H': 13, 'I': 13, 'J': 13
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def show_recap(supabase):
    st.title("📊 Récapitulatifs & Exports Excel Imprimables A4")
    st.markdown("Consultez et téléchargez les synthèses journalières et mensuelles mises en page pour impression **Format A4**.")
    st.markdown("---")

    # Charger les données depuis Supabase
    try:
        res = supabase.table("essai_plaque").select("*").order("date_essai", desc=True).execute()
        data = res.data if res else []
    except Exception as e:
        st.error(f"Erreur lors du chargement des données : {e}")
        return

    if not data:
        st.info("Aucune donnée enregistrée pour le moment.")
        return

    df = pd.DataFrame(data)
    df['date_essai'] = pd.to_datetime(df['date_essai'])

    # --- BARRE DE FILTRES ---
    st.markdown("### 🔍 Filtres de Recherche")
    col_f1, col_f2, col_f3 = st.columns(3)

    with col_f1:
        type_recap = st.selectbox("Type de récapitulatif", ["Journalier", "Mensuel", "Personnalisé par dates"])

    filtered_df = df.copy()
    filter_label = "Général"

    if type_recap == "Journalier":
        with col_f2:
            date_choisie = st.date_input("Choisir le jour", value=date.today())
            filtered_df = df[df['date_essai'].dt.date == date_choisie]
            filter_label = f"Journalier du {date_choisie.strftime('%d/%m/%Y')}"

    elif type_recap == "Mensuel":
        with col_f2:
            mois_annee = st.date_input("Choisir un jour du mois visé", value=date.today())
            filtered_df = df[(df['date_essai'].dt.year == mois_annee.year) & (df['date_essai'].dt.month == mois_annee.month)]
            filter_label = f"Mensuel - {mois_annee.strftime('%m/%Y')}"

    elif type_recap == "Personnalisé par dates":
        with col_f2:
            d_start = st.date_input("Date Début", value=date.today())
        with col_f3:
            d_end = st.date_input("Date Fin", value=date.today())
            filtered_df = df[(df['date_essai'].dt.date >= d_start) & (df['date_essai'].dt.date <= d_end)]
            filter_label = f"Période du {d_start.strftime('%d/%m/%Y')} au {d_end.strftime('%d/%m/%Y')}"

    # Filtre par Emplacement
    with col_f3 if type_recap != "Personnalisé par dates" else col_f1:
        emplacements_dispos = ["Tous les emplacements"] + sorted(list(df['emplacement'].dropna().unique()))
        emp_sel = st.selectbox("Filtrer par Emplacement", emplacements_dispos)
        if emp_sel != "Tous les emplacements":
            filtered_df = filtered_df[filtered_df['emplacement'] == emp_sel]
            filter_label += f" | Emplacement : {emp_sel}"

    st.markdown("---")

    # --- AFFICHAGE DES KPI DANS STREAMLIT ---
    st.markdown(f"### 📈 Synthèse Visuelle : **{filter_label}**")

    if filtered_df.empty:
        st.warning("⚠️ Aucun essai trouvé pour les critères sélectionnés.")
    else:
        k1, k2, k3, k4 = st.columns(4)
        k1.metric("Nombre d'essais", len(filtered_df))
        k2.metric("EV1 Moyen (MPa)", f"{filtered_df['ev1'].mean():.2f}")
        k3.metric("EV2 Moyen (MPa)", f"{filtered_df['ev2'].mean():.2f}")
        k4.metric("Ratio K Moyen", f"{filtered_df['k'].mean():.2f}")

        # Conversion date pour affichage propre
        df_display = filtered_df.copy()
        df_display['date_essai'] = df_display['date_essai'].dt.strftime('%Y-%m-%d')

        st.dataframe(
            df_display[[
                "date_essai", "couche", "emplacement", "pk_profil", 
                "z1", "z2", "ev1", "ev2", "k", "technicien"
            ]].rename(columns={
                "date_essai": "Date", "couche": "Couche", "emplacement": "Emplacement",
                "pk_profil": "PK/Profil", "z1": "Z1 (mm)", "z2": "Z2 (mm)",
                "ev1": "EV1 (MPa)", "ev2": "EV2 (MPa)", "k": "K (EV2/EV1)", "technicien": "Technicien"
            }),
            use_container_width=True,
            hide_index=True
        )

        # --- BOUTON DE TÉLÉCHARGEMENT EXCEL A4 ---
        st.markdown("### 📥 Téléchargement du Rapport Imprimable (A4)")
        
        excel_buffer = generate_excel_recap(df_display, filter_title=filter_label)
        
        file_name_clean = f"Synthese_Essai_Plaque_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        
        st.download_button(
            label="📄 Télécharger la Synthèse Excel (Mise en page A4 Imprimable)",
            data=excel_buffer,
            file_name=file_name_clean,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
